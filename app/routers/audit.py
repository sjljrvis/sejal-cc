# app/routers/audit.py
"""
Audit Log API Endpoints

=============================================================================
AUDIT SYSTEM OVERVIEW
=============================================================================

This module provides read-only API access to audit logs for administrators.
Audit logs are IMMUTABLE - they can never be modified or deleted.

TWO TYPES OF LOGS:
1. Middleware Logs (is_middleware=True)
   - Automatically generated for every HTTP request
   - Contains: method, endpoint, status code, response time, request/response bodies
   - Filtered to exclude health checks and the audit endpoint itself
   - Sensitive data (passwords, tokens) is masked

2. Custom Logs (is_middleware=False)
   - Manually logged via `from app.services.audit import audit`
   - Contains: rich business context ("Admin approved user john@example.com")
   - Used for important business events

ENDPOINTS:
- GET /api/admin/audit - List all audit logs (paginated, filterable)
- GET /api/admin/audit/stats - Get audit statistics
- GET /api/admin/audit/export - Export logs as CSV or Excel
- GET /api/admin/audit/categories - List unique categories
- GET /api/admin/audit/actions - List unique actions
- GET /api/admin/audit/{id} - Get single log by ID
- GET /api/admin/audit/resource/{type}/{id} - Get logs for a resource
- GET /api/admin/audit/actor/{email} - Get logs for an actor

EXPORT FORMATS:
- CSV: /api/admin/audit/export?format=csv
- Excel: /api/admin/audit/export?format=xlsx

=============================================================================
"""

import csv
import io
from datetime import datetime, timedelta
from math import ceil
from typing import Literal, Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import desc, func
from sqlalchemy.orm import Session

from ..core.database import get_db
from ..models.audit import AuditLog
from ..schemas.audit import AuditLogListResponse, AuditLogResponse, AuditStatsResponse
from ..security import get_current_user

router = APIRouter(prefix="/admin/audit", tags=["Audit Logs"])


def _apply_audit_filters(query, filters: dict):
    """Apply filters to an audit log query. Reused for listing and export."""
    if filters.get("category"):
        query = query.filter(AuditLog.category == filters["category"])
    if filters.get("action"):
        query = query.filter(AuditLog.action.ilike(f"%{filters['action']}%"))
    if filters.get("actor_email"):
        query = query.filter(AuditLog.actor_email.ilike(f"%{filters['actor_email']}%"))
    if filters.get("resource_type"):
        query = query.filter(AuditLog.resource_type == filters["resource_type"])
    if filters.get("resource_id"):
        query = query.filter(AuditLog.resource_id == filters["resource_id"])
    if filters.get("success") is not None:
        query = query.filter(AuditLog.success == ("true" if filters["success"] else "false"))
    if filters.get("severity"):
        query = query.filter(AuditLog.severity == filters["severity"])
    if filters.get("start_date"):
        query = query.filter(AuditLog.timestamp >= filters["start_date"])
    if filters.get("end_date"):
        query = query.filter(AuditLog.timestamp <= filters["end_date"])
    if filters.get("search"):
        query = query.filter(AuditLog.description.ilike(f"%{filters['search']}%"))
    # Middleware-specific filters
    if filters.get("http_method"):
        query = query.filter(AuditLog.http_method == filters["http_method"])
    if filters.get("response_status"):
        query = query.filter(AuditLog.response_status == filters["response_status"])
    if filters.get("is_middleware") is not None:
        query = query.filter(AuditLog.is_middleware == filters["is_middleware"])
    return query


@router.get("", response_model=AuditLogListResponse)
async def list_audit_logs(
    # Pagination
    page: int = Query(1, ge=1, description="Page number"),
    page_size: int = Query(50, ge=1, le=100, description="Items per page"),
    # Filters
    category: Optional[str] = Query(None, description="Filter by category"),
    action: Optional[str] = Query(None, description="Filter by action"),
    actor_email: Optional[str] = Query(None, description="Filter by actor email"),
    resource_type: Optional[str] = Query(None, description="Filter by resource type"),
    resource_id: Optional[str] = Query(None, description="Filter by resource ID"),
    success: Optional[bool] = Query(None, description="Filter by success status"),
    severity: Optional[str] = Query(None, description="Filter by severity"),
    start_date: Optional[datetime] = Query(None, description="Start date (ISO format)"),
    end_date: Optional[datetime] = Query(None, description="End date (ISO format)"),
    search: Optional[str] = Query(None, description="Search in description"),
    # Middleware-specific filters
    http_method: Optional[str] = Query(None, description="Filter by HTTP method (GET, POST, etc.)"),
    response_status: Optional[int] = Query(None, description="Filter by HTTP response status"),
    is_middleware: Optional[bool] = Query(None, description="Filter by log source (True=auto, False=custom)"),
    # Dependencies
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    List audit logs with filtering and pagination.
    
    Supports filtering by:
    - Standard fields: category, action, actor, resource, success, severity, date range
    - Middleware fields: HTTP method, response status, is_middleware flag
    
    Requires 'admin' role.
    """
    query = db.query(AuditLog)

    # Build filters dict
    filters = {
        "category": category,
        "action": action,
        "actor_email": actor_email,
        "resource_type": resource_type,
        "resource_id": resource_id,
        "success": success,
        "severity": severity,
        "start_date": start_date,
        "end_date": end_date,
        "search": search,
        "http_method": http_method,
        "response_status": response_status,
        "is_middleware": is_middleware,
    }

    query = _apply_audit_filters(query, filters)

    # Get total count
    total = query.count()

    # Apply pagination and ordering (newest first)
    offset = (page - 1) * page_size
    logs = query.order_by(desc(AuditLog.timestamp)).offset(offset).limit(page_size).all()

    # Calculate total pages
    total_pages = ceil(total / page_size) if total > 0 else 1

    return AuditLogListResponse(
        logs=[AuditLogResponse.model_validate(log) for log in logs],
        total=total,
        page=page,
        page_size=page_size,
        total_pages=total_pages,
    )


@router.get("/stats", response_model=AuditStatsResponse)
async def get_audit_stats(
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get comprehensive audit log statistics.
    
    Returns:
    - Total events, events today/this week
    - Breakdown by category and action
    - Recent errors
    - Middleware-specific: by HTTP method, by status code, avg response time
    - Split between middleware (auto) and custom logs
    
    Requires 'admin' role.
    """
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)

    # Total events
    total_events = db.query(func.count(AuditLog.id)).scalar() or 0

    # Events today
    events_today = (
        db.query(func.count(AuditLog.id))
        .filter(AuditLog.timestamp >= today_start)
        .scalar()
        or 0
    )

    # Events this week
    events_this_week = (
        db.query(func.count(AuditLog.id))
        .filter(AuditLog.timestamp >= week_start)
        .scalar()
        or 0
    )

    # By category
    category_counts = (
        db.query(AuditLog.category, func.count(AuditLog.id))
        .group_by(AuditLog.category)
        .all()
    )
    by_category = {cat: count for cat, count in category_counts}

    # By action (top 10)
    action_counts = (
        db.query(AuditLog.action, func.count(AuditLog.id))
        .group_by(AuditLog.action)
        .order_by(desc(func.count(AuditLog.id)))
        .limit(10)
        .all()
    )
    by_action = {action: count for action, count in action_counts}

    # Recent errors (last 7 days)
    recent_errors = (
        db.query(func.count(AuditLog.id))
        .filter(AuditLog.success == "false")
        .filter(AuditLog.timestamp >= week_start)
        .scalar()
        or 0
    )

    # Middleware-specific stats
    # By HTTP method
    method_counts = (
        db.query(AuditLog.http_method, func.count(AuditLog.id))
        .filter(AuditLog.http_method.isnot(None))
        .group_by(AuditLog.http_method)
        .all()
    )
    by_http_method = {method: count for method, count in method_counts if method}

    # By response status (grouped by first digit: 2xx, 4xx, 5xx)
    status_counts = (
        db.query(AuditLog.response_status, func.count(AuditLog.id))
        .filter(AuditLog.response_status.isnot(None))
        .group_by(AuditLog.response_status)
        .all()
    )
    by_status_code = {}
    for status, count in status_counts:
        if status:
            key = f"{status // 100}xx"
            by_status_code[key] = by_status_code.get(key, 0) + count

    # Average response time
    avg_response_time = (
        db.query(func.avg(AuditLog.response_time_ms))
        .filter(AuditLog.response_time_ms.isnot(None))
        .scalar()
    )

    # Middleware vs custom log counts
    middleware_logs = (
        db.query(func.count(AuditLog.id))
        .filter(AuditLog.is_middleware == True)
        .scalar()
        or 0
    )
    custom_logs = total_events - middleware_logs

    return AuditStatsResponse(
        total_events=total_events,
        events_today=events_today,
        events_this_week=events_this_week,
        by_category=by_category,
        by_action=by_action,
        recent_errors=recent_errors,
        by_http_method=by_http_method,
        by_status_code=by_status_code,
        avg_response_time_ms=round(avg_response_time, 2) if avg_response_time else None,
        middleware_logs=middleware_logs,
        custom_logs=custom_logs,
    )


@router.get("/export")
async def export_audit_logs(
    # Format
    format: Literal["csv", "xlsx"] = Query("csv", description="Export format: csv or xlsx"),
    # Filters (same as list)
    category: Optional[str] = Query(None),
    action: Optional[str] = Query(None),
    actor_email: Optional[str] = Query(None),
    resource_type: Optional[str] = Query(None),
    success: Optional[bool] = Query(None),
    severity: Optional[str] = Query(None),
    start_date: Optional[datetime] = Query(None),
    end_date: Optional[datetime] = Query(None),
    http_method: Optional[str] = Query(None),
    response_status: Optional[int] = Query(None),
    is_middleware: Optional[bool] = Query(None),
    # Limit
    limit: int = Query(10000, ge=1, le=100000, description="Max rows to export"),
    # Dependencies
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Export audit logs as CSV or Excel.
    
    Supports the same filters as the list endpoint.
    Maximum 100,000 rows per export to prevent memory issues.
    
    Requires 'admin' role.
    """
    query = db.query(AuditLog)

    # Apply filters
    filters = {
        "category": category,
        "action": action,
        "actor_email": actor_email,
        "resource_type": resource_type,
        "success": success,
        "severity": severity,
        "start_date": start_date,
        "end_date": end_date,
        "http_method": http_method,
        "response_status": response_status,
        "is_middleware": is_middleware,
    }
    query = _apply_audit_filters(query, filters)

    # Fetch logs (limited)
    logs = query.order_by(desc(AuditLog.timestamp)).limit(limit).all()

    if format == "csv":
        return _export_csv(logs)
    else:
        return _export_xlsx(logs)


def _export_csv(logs: list[AuditLog]) -> StreamingResponse:
    """Generate CSV export."""
    output = io.StringIO()
    
    # Define columns
    fieldnames = [
        "ID", "Timestamp", "Actor Email", "Actor IP", "Action", "Category",
        "Severity", "Resource Type", "Resource ID", "Description", "Success",
        "Error", "HTTP Method", "Endpoint", "Status Code", "Response Time (ms)", "Source"
    ]
    
    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()
    
    for log in logs:
        writer.writerow(log.to_export_dict())
    
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f"audit_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_xlsx(logs: list[AuditLog]) -> StreamingResponse:
    """
    Generate Excel export.
    
    Note: Requires openpyxl package. Falls back to CSV if not available.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return _export_csv(logs)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"
    
    # Headers with styling
    headers = [
        "ID", "Timestamp", "Actor Email", "Actor IP", "Action", "Category",
        "Severity", "Resource Type", "Resource ID", "Description", "Success",
        "Error", "HTTP Method", "Endpoint", "Status Code", "Response Time (ms)", "Source"
    ]
    
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Data rows
    for row_idx, log in enumerate(logs, 2):
        export_dict = log.to_export_dict()
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=export_dict.get(header, ""))
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"audit_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/categories")
async def list_audit_categories(
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    List all unique audit categories.
    Useful for building filter dropdowns in the UI.
    """
    categories = db.query(AuditLog.category).distinct().all()
    return {"categories": [c[0] for c in categories if c[0]]}


@router.get("/actions")
async def list_audit_actions(
    category: Optional[str] = Query(None, description="Filter by category"),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    List all unique audit actions, optionally filtered by category.
    Useful for building filter dropdowns in the UI.
    """
    query = db.query(AuditLog.action).distinct()
    if category:
        query = query.filter(AuditLog.category == category)
    actions = query.all()
    return {"actions": [a[0] for a in actions if a[0]]}


@router.get("/{audit_id}", response_model=AuditLogResponse)
async def get_audit_log(
    audit_id: int,
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get a specific audit log entry by ID.
    Requires 'admin' role.
    """
    from fastapi import HTTPException

    log = db.query(AuditLog).filter(AuditLog.id == audit_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="Audit log not found")

    return AuditLogResponse.model_validate(log)


@router.get("/resource/{resource_type}/{resource_id}")
async def get_resource_audit_trail(
    resource_type: str,
    resource_id: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=50),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get audit trail for a specific resource.
    Useful for viewing all actions on a specific user, item, etc.
    """
    query = (
        db.query(AuditLog)
        .filter(AuditLog.resource_type == resource_type)
        .filter(AuditLog.resource_id == resource_id)
    )

    total = query.count()
    offset = (page - 1) * page_size
    logs = query.order_by(desc(AuditLog.timestamp)).offset(offset).limit(page_size).all()

    return {
        "resource_type": resource_type,
        "resource_id": resource_id,
        "logs": [AuditLogResponse.model_validate(log) for log in logs],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/actor/{actor_email}")
async def get_actor_audit_trail(
    actor_email: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=50),
    user: dict = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Get audit trail for a specific actor.
    Useful for viewing all actions by a specific admin/user.
    """
    query = db.query(AuditLog).filter(AuditLog.actor_email.ilike(f"%{actor_email}%"))

    total = query.count()
    offset = (page - 1) * page_size
    logs = query.order_by(desc(AuditLog.timestamp)).offset(offset).limit(page_size).all()

    return {
        "actor_email": actor_email,
        "logs": [AuditLogResponse.model_validate(log) for log in logs],
        "total": total,
        "page": page,
        "page_size": page_size,
    }

